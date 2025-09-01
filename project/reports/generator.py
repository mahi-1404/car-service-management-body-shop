import os
from datetime import datetime, date, timedelta
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from models.inventory import Inventory
from models.photos import Photo
from models.work_status import WorkStatus, WorkItem
from models.claims import Claim
from models.approvals import Approval
from models.registration_status import RegistrationStatus

class ReportGenerator:
    """Generate various reports for the car service management system"""
    
    def __init__(self):
        self.reports_dir = 'reports'
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_daily_report(self, format_type='pdf'):
        """Generate daily report"""
        today = date.today()
        vehicles = Inventory.query.filter(
            Inventory.check_in_date >= datetime.combine(today, datetime.min.time()),
            Inventory.check_in_date < datetime.combine(today + timedelta(days=1), datetime.min.time())
        ).all()
        
        filename = f"daily_report_{today.strftime('%Y%m%d')}"
        
        if format_type == 'pdf':
            return self._generate_pdf_report(vehicles, f"Daily Report - {today.strftime('%B %d, %Y')}", filename)
        else:
            return self._generate_excel_report(vehicles, f"Daily Report - {today.strftime('%B %d, %Y')}", filename)
    
    def generate_monthly_report(self, format_type='pdf'):
        """Generate monthly report"""
        today = date.today()
        first_day = today.replace(day=1)
        vehicles = Inventory.query.filter(
            Inventory.check_in_date >= datetime.combine(first_day, datetime.min.time()),
            Inventory.check_in_date < datetime.combine(today + timedelta(days=1), datetime.min.time())
        ).all()
        
        filename = f"monthly_report_{today.strftime('%Y%m')}"
        
        if format_type == 'pdf':
            return self._generate_pdf_report(vehicles, f"Monthly Report - {today.strftime('%B %Y')}", filename)
        else:
            return self._generate_excel_report(vehicles, f"Monthly Report - {today.strftime('%B %Y')}", filename)

    def generate_all_cars_report(self, format_type='pdf'):
        """Generate a report of all cars"""
        vehicles = Inventory.query.order_by(Inventory.serial_number.asc()).all()
        filename = f"all_cars_report_{datetime.now().strftime('%Y%m%d')}"
        
        if format_type == 'pdf':
            return self._generate_all_cars_pdf_report(vehicles, "All Cars Report", filename)
        else:
            # Optionally, implement Excel generation for this report as well
            return None

    def _generate_all_cars_pdf_report(self, vehicles, title, filename):
        """Generate PDF report for all cars"""
        filepath = os.path.join(self.reports_dir, f"{filename}.pdf")
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1
        )
        
        content = []
        content.append(Paragraph(title, title_style))
        
        data = [['S.No', 'Vehicle No.', 'Customer', 'Phone', 'Insurance', 'Claim No.', 'Engine No.', 'Chassis No.', 'Check-in']]
        
        for vehicle in vehicles:
            claim = Claim.query.filter_by(vehicle_number=vehicle.vehicle_number).first()
            data.append([
                str(vehicle.serial_number),
                vehicle.vehicle_number,
                vehicle.customer_name,
                vehicle.phone_number,
                vehicle.insurance_name,
                claim.claim_number if claim else 'N/A',
                vehicle.engine_number,
                vehicle.chassis_number,
                vehicle.check_in_date.strftime('%Y-%m-%d %H:%M')
            ])

        table = Table(data, colWidths=[0.4*inch, 0.9*inch, 1.2*inch, 0.9*inch, 1.2*inch, 0.8*inch, 0.9*inch, 1.1*inch, 1.1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        content.append(table)
        doc.build(content)
        return filepath
    
    def _generate_pdf_report(self, vehicles, title, filename):
        """Generate PDF report"""
        filepath = os.path.join(self.reports_dir, f"{filename}.pdf")
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Content
        content = []
        
        # Title
        content.append(Paragraph(title, title_style))
        content.append(Spacer(1, 20))
        
        # Summary
        total_vehicles = len(vehicles)
        completed_vehicles = sum(1 for v in vehicles if self._is_vehicle_completed(v))
        
        summary_data = [
            ['Total Vehicles', str(total_vehicles)],
            ['Completed Vehicles', str(completed_vehicles)],
            ['Pending Vehicles', str(total_vehicles - completed_vehicles)],
            ['Completion Rate', f"{(completed_vehicles/total_vehicles*100):.1f}%" if total_vehicles > 0 else "0%"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        content.append(summary_table)
        content.append(Spacer(1, 30))
        
        # Vehicle details
        if vehicles:
            content.append(Paragraph("Vehicle Details", styles['Heading2']))
            content.append(Spacer(1, 10))
            
            # Table headers
            data = [['S.No', 'Vehicle No.', 'Check-in Date', 'Status', 'Progress']]
            
            for vehicle in vehicles:
                progress = self._calculate_vehicle_progress(vehicle)
                status = 'Completed' if progress == 100 else 'In Progress'
                
                data.append([
                    str(vehicle.serial_number),
                    vehicle.vehicle_number,
                    vehicle.check_in_date.strftime('%Y-%m-%d'),
                    status,
                    f"{progress}%"
                ])
            
            table = Table(data, colWidths=[0.8*inch, 1.5*inch, 1.5*inch, 1.2*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            content.append(table)
        
        # Build PDF
        doc.build(content)
        return filepath
    
    def _generate_excel_report(self, vehicles, title, filename):
        """Generate Excel report"""
        filepath = os.path.join(self.reports_dir, f"{filename}.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Report"
        
        # Title
        sheet['A1'] = title
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A1:E1')
        
        # Summary
        sheet['A3'] = 'Summary'
        sheet['A3'].font = Font(size=14, bold=True)
        
        total_vehicles = len(vehicles)
        completed_vehicles = sum(1 for v in vehicles if self._is_vehicle_completed(v))
        
        sheet['A4'] = 'Total Vehicles'
        sheet['B4'] = total_vehicles
        sheet['A5'] = 'Completed Vehicles'
        sheet['B5'] = completed_vehicles
        sheet['A6'] = 'Pending Vehicles'
        sheet['B6'] = total_vehicles - completed_vehicles
        sheet['A7'] = 'Completion Rate'
        sheet['B7'] = f"{(completed_vehicles/total_vehicles*100):.1f}%" if total_vehicles > 0 else "0%"
        
        # Headers
        headers = ['S.No', 'Vehicle No.', 'Check-in Date', 'Status', 'Progress']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=9, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        for row, vehicle in enumerate(vehicles, 10):
            progress = self._calculate_vehicle_progress(vehicle)
            status = 'Completed' if progress == 100 else 'In Progress'
            
            sheet.cell(row=row, column=1).value = vehicle.serial_number
            sheet.cell(row=row, column=2).value = vehicle.vehicle_number
            sheet.cell(row=row, column=3).value = vehicle.check_in_date.strftime('%Y-%m-%d')
            sheet.cell(row=row, column=4).value = status
            sheet.cell(row=row, column=5).value = f"{progress}%"
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(filepath)
        return filepath
    
    def _is_vehicle_completed(self, vehicle):
        """Check if a vehicle has completed all steps"""
        # Check registration
        registration = RegistrationStatus.query.filter_by(vehicle_number=vehicle.vehicle_number).first()
        if not (registration and registration.is_completed):
            return False
        
        # Check claim
        claim = Claim.query.filter_by(vehicle_number=vehicle.vehicle_number).first()
        if not (claim and claim.claim_number):
            return False
        
        # Check approval
        approval = Approval.query.filter_by(vehicle_number=vehicle.vehicle_number).first()
        if not (approval and approval.is_approved):
            return False
        
        # Check work status
        work_status = WorkStatus.query.filter_by(vehicle_number=vehicle.vehicle_number).first()
        if not work_status:
            return False
        
        work_items = WorkItem.query.filter_by(work_status_id=work_status.id).all()
        if not work_items or not all(item.is_completed for item in work_items):
            return False
        
        return True
    
    def _calculate_vehicle_progress(self, vehicle):
        """Calculate completion percentage for a vehicle"""
        total_steps = 5
        completed_steps = 1  # Inventory is always completed
        
        # Check photos
        if Photo.query.filter_by(vehicle_number=vehicle.vehicle_number).first():
            completed_steps += 1
        
        # Check registration
        registration = RegistrationStatus.query.filter_by(vehicle_number=vehicle.vehicle_number).first()
        if registration and registration.is_completed:
            completed_steps += 1
        
        # Check claim
        claim = Claim.query.filter_by(vehicle_number=vehicle.vehicle_number).first()
        if claim and claim.claim_number:
            completed_steps += 1
        
        # Check work status
        work_status = WorkStatus.query.filter_by(vehicle_number=vehicle.vehicle_number).first()
        if work_status:
            work_items = WorkItem.query.filter_by(work_status_id=work_status.id).all()
            if work_items and all(item.is_completed for item in work_items):
                completed_steps += 1
        
        return int((completed_steps / total_steps) * 100)
